from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File, Form, BackgroundTasks
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import FileResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import jwt
from passlib.context import CryptContext
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
import json
import ftplib
import asyncio
from contextlib import asynccontextmanager

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Exports directory for Excel files
EXPORTS_DIR = ROOT_DIR / "exports"
EXPORTS_DIR.mkdir(exist_ok=True)

# FTP Configuration
FTP_HOST = os.environ.get('FTP_HOST', '')
FTP_USER = os.environ.get('FTP_USER', '')
FTP_PASS = os.environ.get('FTP_PASS', '')
FTP_FOLDER = os.environ.get('FTP_FOLDER', '/dolgozok_backup')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Configuration
JWT_SECRET = os.environ.get('JWT_SECRET', 'dolgozocrm-secret-key-2024')
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# Security
security = HTTPBearer()

app = FastAPI(title="Dolgozó CRM API")
api_router = APIRouter(prefix="/api")

# ==================== MODELS ====================
from fastapi.middleware.cors import CORSMiddleware

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://dolgozocrm.optimalcrew.hu",
        "http://dolgozocrm.optimalcrew.hu",
        "https://dolgozocrmfinal.onrender.com",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: Optional[str] = ""
    role: str = "user"  # "admin" or "user" (toborzó)

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserResponse(BaseModel):
    id: str
    email: str
    name: str
    role: str
    created_at: str

class PasswordChange(BaseModel):
    current_password: str
    new_password: str

class ProfileUpdate(BaseModel):
    name: str

class WorkerTypeCreate(BaseModel):
    name: str

class WorkerTypeResponse(BaseModel):
    id: str
    name: str

class CategoryCreate(BaseModel):
    name: str
    color: str = "#3b82f6"  # Default blue

class CategoryUpdate(BaseModel):
    name: Optional[str] = None
    color: Optional[str] = None
    order: Optional[int] = None

class CategoryResponse(BaseModel):
    id: str
    name: str
    color: str = "#3b82f6"
    order: int = 0
    worker_count: int = 0

class PositionCreate(BaseModel):
    name: str
    worker_type_id: str  # Melyik típushoz tartozik

class PositionResponse(BaseModel):
    id: str
    name: str
    worker_type_id: str
    worker_type_name: Optional[str] = ""

class StatusCreate(BaseModel):
    name: str

class StatusResponse(BaseModel):
    id: str
    name: str

class TagCreate(BaseModel):
    name: str
    color: str = "#6366f1"

class TagResponse(BaseModel):
    id: str
    name: str
    color: str

class WorkerCreate(BaseModel):
    name: str
    phone: str
    worker_type_id: str
    position: Optional[str] = ""  # Szabad szöveg pozíció
    position_experience: Optional[str] = ""  # Pozícióval kapcsolatos tapasztalat
    category: str = "Felvitt dolgozók"
    address: Optional[str] = ""
    email: Optional[str] = ""
    experience: Optional[str] = ""
    notes: Optional[str] = ""
    global_status: str = "Feldolgozatlan"  # Alap (globális) dolgozói státusz
    project_id: Optional[str] = None  # Opcionális projekt várólistához
    start_date: Optional[str] = None  # Tervezett kezdési dátum (ha project_id megadott)

class WorkerUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    worker_type_id: Optional[str] = None
    position: Optional[str] = None
    position_experience: Optional[str] = None
    category: Optional[str] = None
    address: Optional[str] = None
    email: Optional[str] = None
    experience: Optional[str] = None
    notes: Optional[str] = None
    global_status: Optional[str] = None

class WorkerResponse(BaseModel):
    id: str
    name: str
    phone: str
    worker_type_id: str
    worker_type_name: Optional[str] = ""
    position: Optional[str] = ""
    position_experience: Optional[str] = ""
    category: str
    address: str
    email: str
    experience: str
    notes: str
    global_status: str = "Feldolgozatlan"
    tags: List[dict] = []
    project_statuses: List[dict] = []
    owner_id: str
    owner_name: str
    created_at: str

class ProjectCreate(BaseModel):
    name: str
    client_name: Optional[str] = ""  # Ügyfél / cégnév
    date: str
    location: Optional[str] = ""  # Helyszín
    training_location: Optional[str] = ""  # Betanítás / munkavégzés helye
    notes: Optional[str] = ""  # Megjegyzések, elvárások
    recruiter_ids: List[str] = []  # Hozzárendelt toborzók

class ProjectUpdate(BaseModel):
    name: Optional[str] = None
    client_name: Optional[str] = None
    date: Optional[str] = None
    location: Optional[str] = None
    training_location: Optional[str] = None
    notes: Optional[str] = None
    is_closed: Optional[bool] = None
    recruiter_ids: Optional[List[str]] = None

class ProjectResponse(BaseModel):
    id: str
    name: str
    client_name: str = ""
    date: str
    location: str
    training_location: str = ""
    notes: str
    is_closed: bool
    worker_count: int
    position_count: int = 0
    total_headcount: int = 0  # Összesített létszámigény pozíciókból
    trial_count: int = 0
    recruiter_ids: List[str] = []
    recruiters: List[dict] = []
    owner_id: str = ""
    owner_name: str = ""
    created_at: str

# ==================== PROJECT POSITION MODELS ====================

class ProjectPositionCreate(BaseModel):
    name: str  # Pozíció neve (pl. Operátor, Raktáros)
    headcount: int = 1  # Létszámigény
    shift_schedule: Optional[str] = ""  # Műszakrend
    experience_required: Optional[str] = ""  # Tapasztalat
    qualifications: Optional[str] = ""  # Végzettség / jogosítvány
    physical_requirements: Optional[str] = ""  # Fizikai elvárások
    notes: Optional[str] = ""  # Egyéb megjegyzések

class ProjectPositionUpdate(BaseModel):
    name: Optional[str] = None
    headcount: Optional[int] = None
    shift_schedule: Optional[str] = None
    experience_required: Optional[str] = None
    qualifications: Optional[str] = None
    physical_requirements: Optional[str] = None
    notes: Optional[str] = None

class ProjectPositionResponse(BaseModel):
    id: str
    project_id: str
    name: str
    headcount: int
    shift_schedule: str = ""
    experience_required: str = ""
    qualifications: str = ""
    physical_requirements: str = ""
    notes: str = ""
    assigned_workers: int = 0
    created_at: str

# ==================== TRIAL MODELS ====================

class TrialCreate(BaseModel):
    date: str  # Próba dátuma
    time: Optional[str] = ""  # Próba időpontja (pl. "09:00")
    notes: Optional[str] = ""

class TrialUpdate(BaseModel):
    date: Optional[str] = None
    time: Optional[str] = None
    notes: Optional[str] = None

class TrialPositionCreate(BaseModel):
    position_id: Optional[str] = None  # Meglévő projekt pozíció ID
    position_name: str  # Pozíció neve (új vagy meglévő)
    headcount: int = 1  # Létszámigény
    hourly_rate: Optional[str] = ""  # Órabér
    accommodation: bool = False  # Van-e szállás
    requirements: str = ""  # Egyéb elvárások
    add_to_project: bool = False  # Új pozíció esetén hozzáadjuk a projekthez?

class TrialPositionResponse(BaseModel):
    id: str
    trial_id: str
    position_id: Optional[str] = None
    position_name: str
    headcount: int = 1
    hourly_rate: str = ""
    accommodation: bool = False
    requirements: str = ""
    assigned_count: int = 0  # Hány dolgozó van hozzárendelve

class TrialResponse(BaseModel):
    id: str
    project_id: str
    date: str
    time: str = ""
    notes: str = ""
    worker_count: int = 0
    workers: List[dict] = []
    positions: List[TrialPositionResponse] = []  # Próba pozíciók
    created_at: str

class TrialWorkerAdd(BaseModel):
    worker_id: str
    position_id: Optional[str] = None  # Melyik próba pozícióra (trial_position_id)

class ProjectWorkerAdd(BaseModel):
    worker_id: str
    status_id: Optional[str] = None

class ProjectRecruiterAdd(BaseModel):
    user_id: str

class ProjectWorkerStatusUpdate(BaseModel):
    status_id: str
    notes: Optional[str] = None

class WorkerHistoryEntry(BaseModel):
    project_id: str
    project_name: str
    project_date: str
    status_id: str
    status_name: str
    notes: str
    updated_at: str

# ==================== WAITLIST MODELS ====================

class WaitlistWorkerAdd(BaseModel):
    worker_id: str
    start_date: Optional[str] = None  # Tervezett munkakezdési dátum
    notes: Optional[str] = ""

class WaitlistWorkerUpdate(BaseModel):
    start_date: Optional[str] = None
    notes: Optional[str] = None

class WaitlistWorkerResponse(BaseModel):
    id: str
    project_id: str
    worker_id: str
    worker_name: str
    worker_phone: str
    worker_email: str = ""
    start_date: str = ""
    notes: str = ""
    added_at: str
    added_by: str
    added_by_name: str = ""

# ==================== NOTIFICATION MODELS ====================

class NotificationResponse(BaseModel):
    id: str
    user_id: str
    type: str  # "project_assigned", "trial_assigned"
    title: str
    message: str
    link: str = ""
    is_read: bool = False
    created_at: str

# ==================== HELPER FUNCTIONS ====================

def hash_password(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)

def create_token(user_id: str, email: str, role: str) -> str:
    payload = {
        "user_id": user_id,
        "email": email,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def create_notification(user_id: str, notification_type: str, title: str, message: str, link: str = ""):
    """Helper function to create a notification"""
    notif_doc = {
        "id": str(uuid.uuid4()),
        "user_id": user_id,
        "type": notification_type,
        "title": title,
        "message": message,
        "link": link,
        "is_read": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.notifications.insert_one(notif_doc)
    return notif_doc

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"id": payload["user_id"]}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="Felhasználó nem található")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token lejárt")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Érvénytelen token")

async def require_admin(user: dict = Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Csak admin jogosultsággal")
    return user

# ==================== AUTH ENDPOINTS ====================

@api_router.post("/auth/register", response_model=dict)
async def register(data: UserCreate, current_user: dict = Depends(require_admin)):
    """Admin csak hozhat létre új felhasználót"""
    existing = await db.users.find_one({"email": data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Ez az email már regisztrálva van")
    
    if len(data.password) < 8:
        raise HTTPException(status_code=400, detail="A jelszó minimum 8 karakter legyen")
    
    user_doc = {
        "id": str(uuid.uuid4()),
        "email": data.email,
        "password": hash_password(data.password),
        "name": data.name or data.email.split("@")[0],
        "role": data.role,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    return {"message": "Felhasználó létrehozva", "email": data.email}

@api_router.post("/auth/login", response_model=dict)
async def login(data: UserLogin):
    user = await db.users.find_one({"email": data.email}, {"_id": 0})
    if not user or not verify_password(data.password, user["password"]):
        raise HTTPException(status_code=401, detail="Hibás email vagy jelszó")
    
    token = create_token(user["id"], user["email"], user["role"])
    return {
        "token": token,
        "user": {
            "id": user["id"],
            "email": user["email"],
            "name": user["name"],
            "role": user["role"]
        }
    }

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(user: dict = Depends(get_current_user)):
    return UserResponse(
        id=user["id"],
        email=user["email"],
        name=user.get("name", ""),
        role=user["role"],
        created_at=user.get("created_at", "")
    )

@api_router.put("/auth/profile")
async def update_profile(data: ProfileUpdate, user: dict = Depends(get_current_user)):
    """Csak admin tudja módosítani a profilt"""
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Csak admin módosíthatja a profilt")
    await db.users.update_one({"id": user["id"]}, {"$set": {"name": data.name}})
    return {"message": "Profil frissítve"}

@api_router.put("/auth/password")
async def change_password(data: PasswordChange, user: dict = Depends(get_current_user)):
    db_user = await db.users.find_one({"id": user["id"]}, {"_id": 0})
    if not verify_password(data.current_password, db_user["password"]):
        raise HTTPException(status_code=400, detail="Hibás jelenlegi jelszó")
    
    if len(data.new_password) < 8:
        raise HTTPException(status_code=400, detail="Az új jelszó minimum 8 karakter legyen")
    
    await db.users.update_one(
        {"id": user["id"]}, 
        {"$set": {"password": hash_password(data.new_password)}}
    )
    return {"message": "Jelszó megváltoztatva"}

# ==================== WORKER TYPES ====================

@api_router.get("/worker-types", response_model=List[WorkerTypeResponse])
async def get_worker_types(user: dict = Depends(get_current_user)):
    types = await db.worker_types.find({}, {"_id": 0}).to_list(100)
    return [WorkerTypeResponse(**t) for t in types]

@api_router.post("/worker-types", response_model=WorkerTypeResponse)
async def create_worker_type(data: WorkerTypeCreate, user: dict = Depends(require_admin)):
    type_doc = {"id": str(uuid.uuid4()), "name": data.name}
    await db.worker_types.insert_one(type_doc)
    return WorkerTypeResponse(**type_doc)

@api_router.delete("/worker-types/{type_id}")
async def delete_worker_type(type_id: str, user: dict = Depends(require_admin)):
    result = await db.worker_types.delete_one({"id": type_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Típus nem található")
    # Töröljük a típushoz tartozó pozíciókat is
    await db.positions.delete_many({"worker_type_id": type_id})
    return {"message": "Típus törölve"}

# ==================== POSITIONS ====================

@api_router.get("/positions", response_model=List[PositionResponse])
async def get_positions(worker_type_id: Optional[str] = None, user: dict = Depends(get_current_user)):
    """Pozíciók lekérése, opcionálisan típus szerint szűrve"""
    query = {}
    if worker_type_id:
        query["worker_type_id"] = worker_type_id
    
    positions = await db.positions.find(query, {"_id": 0}).to_list(100)
    
    result = []
    for p in positions:
        type_doc = await db.worker_types.find_one({"id": p.get("worker_type_id")}, {"_id": 0})
        result.append(PositionResponse(
            id=p["id"],
            name=p["name"],
            worker_type_id=p["worker_type_id"],
            worker_type_name=type_doc["name"] if type_doc else ""
        ))
    return result

@api_router.post("/positions", response_model=PositionResponse)
async def create_position(data: PositionCreate, user: dict = Depends(require_admin)):
    # Ellenőrizzük, hogy létezik-e a típus
    type_doc = await db.worker_types.find_one({"id": data.worker_type_id}, {"_id": 0})
    if not type_doc:
        raise HTTPException(status_code=404, detail="Típus nem található")
    
    position_doc = {
        "id": str(uuid.uuid4()),
        "name": data.name,
        "worker_type_id": data.worker_type_id
    }
    await db.positions.insert_one(position_doc)
    return PositionResponse(**position_doc, worker_type_name=type_doc["name"])

@api_router.delete("/positions/{position_id}")
async def delete_position(position_id: str, user: dict = Depends(require_admin)):
    result = await db.positions.delete_one({"id": position_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Pozíció nem található")
    return {"message": "Pozíció törölve"}

# ==================== STATUSES ====================

@api_router.get("/statuses", response_model=List[StatusResponse])
async def get_statuses(user: dict = Depends(get_current_user)):
    statuses = await db.statuses.find({}, {"_id": 0}).to_list(100)
    return [StatusResponse(**s) for s in statuses]

@api_router.post("/statuses", response_model=StatusResponse)
async def create_status(data: StatusCreate, user: dict = Depends(require_admin)):
    status_doc = {"id": str(uuid.uuid4()), "name": data.name}
    await db.statuses.insert_one(status_doc)
    return StatusResponse(**status_doc)

@api_router.delete("/statuses/{status_id}")
async def delete_status(status_id: str, user: dict = Depends(require_admin)):
    result = await db.statuses.delete_one({"id": status_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Státusz nem található")
    return {"message": "Státusz törölve"}

# ==================== TAGS ====================

@api_router.get("/tags", response_model=List[TagResponse])
async def get_tags(user: dict = Depends(get_current_user)):
    tags = await db.tags.find({}, {"_id": 0}).to_list(100)
    return [TagResponse(**t) for t in tags]

@api_router.post("/tags", response_model=TagResponse)
async def create_tag(data: TagCreate, user: dict = Depends(require_admin)):
    tag_doc = {"id": str(uuid.uuid4()), "name": data.name, "color": data.color}
    await db.tags.insert_one(tag_doc)
    return TagResponse(**tag_doc)

@api_router.delete("/tags/{tag_id}")
async def delete_tag(tag_id: str, user: dict = Depends(require_admin)):
    result = await db.tags.delete_one({"id": tag_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Jellemző nem található")
    return {"message": "Jellemző törölve"}

# ==================== CATEGORIES ====================

@api_router.get("/categories", response_model=List[CategoryResponse])
async def get_categories(user: dict = Depends(get_current_user)):
    """Kategóriák lekérése worker count-tal és rendezve"""
    categories = await db.categories.find({}, {"_id": 0}).sort("order", 1).to_list(100)
    # Ha nincs még kategória, visszaadjuk az alapértelmezetteket
    if not categories:
        default_categories = [
            {"id": str(uuid.uuid4()), "name": "Felvitt dolgozók", "color": "#3b82f6", "order": 0},
            {"id": str(uuid.uuid4()), "name": "Hideg jelentkező", "color": "#22c55e", "order": 1},
            {"id": str(uuid.uuid4()), "name": "Űrlapon jelentkezett", "color": "#f97316", "order": 2},
            {"id": str(uuid.uuid4()), "name": "Állásra jelentkezett", "color": "#a855f7", "order": 3},
            {"id": str(uuid.uuid4()), "name": "Ingázó", "color": "#64748b", "order": 4},
            {"id": str(uuid.uuid4()), "name": "Szállásos", "color": "#f59e0b", "order": 5},
        ]
        await db.categories.insert_many(default_categories)
        categories = default_categories
    
    # Add worker count to each category
    result = []
    for c in categories:
        worker_count = await db.workers.count_documents({"category": c["name"]})
        result.append(CategoryResponse(
            id=c["id"],
            name=c["name"],
            color=c.get("color", "#3b82f6"),
            order=c.get("order", 0),
            worker_count=worker_count
        ))
    return result

@api_router.get("/categories/stats")
async def get_category_stats(user: dict = Depends(get_current_user)):
    """Kategória statisztikák dashboard-hoz"""
    categories = await db.categories.find({}, {"_id": 0}).sort("order", 1).to_list(100)
    
    stats = []
    total_workers = 0
    for cat in categories:
        count = await db.workers.count_documents({"category": cat["name"]})
        total_workers += count
        stats.append({
            "id": cat["id"],
            "name": cat["name"],
            "color": cat.get("color", "#3b82f6"),
            "count": count
        })
    
    # Recent activity - workers added in last 7 days per category
    from datetime import timedelta
    week_ago = (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()
    
    recent_stats = []
    for cat in categories:
        recent_count = await db.workers.count_documents({
            "category": cat["name"],
            "created_at": {"$gte": week_ago}
        })
        recent_stats.append({
            "name": cat["name"],
            "color": cat.get("color", "#3b82f6"),
            "count": recent_count
        })
    
    return {
        "total_workers": total_workers,
        "category_stats": stats,
        "recent_activity": recent_stats,
        "categories_count": len(categories)
    }

@api_router.post("/categories", response_model=CategoryResponse)
async def create_category(data: CategoryCreate, user: dict = Depends(require_admin)):
    """Új kategória létrehozása - csak admin"""
    # Ellenőrizzük, hogy ne legyen duplikált név
    existing = await db.categories.find_one({"name": data.name})
    if existing:
        raise HTTPException(status_code=400, detail="Ilyen nevű kategória már létezik")
    
    # Get max order
    max_order_cat = await db.categories.find_one({}, {"_id": 0}, sort=[("order", -1)])
    next_order = (max_order_cat.get("order", 0) + 1) if max_order_cat else 0
    
    category_doc = {
        "id": str(uuid.uuid4()),
        "name": data.name,
        "color": data.color,
        "order": next_order
    }
    await db.categories.insert_one(category_doc)
    return CategoryResponse(**category_doc, worker_count=0)

class CategoryOrderItem(BaseModel):
    id: str
    order: int

class CategoryReorderRequest(BaseModel):
    orders: List[CategoryOrderItem]

@api_router.put("/categories/reorder")
async def reorder_categories(data: CategoryReorderRequest, user: dict = Depends(require_admin)):
    """Kategóriák átrendezése - csak admin"""
    for item in data.orders:
        await db.categories.update_one(
            {"id": item.id},
            {"$set": {"order": item.order}}
        )
    return {"message": "Sorrend frissítve"}

@api_router.put("/categories/{category_id}", response_model=CategoryResponse)
async def update_category(category_id: str, data: CategoryUpdate, user: dict = Depends(require_admin)):
    """Kategória szerkesztése - csak admin"""
    category = await db.categories.find_one({"id": category_id}, {"_id": 0})
    if not category:
        raise HTTPException(status_code=404, detail="Kategória nem található")
    
    update_data = {}
    old_name = category["name"]
    
    if data.name is not None and data.name != old_name:
        # Check for duplicate name
        existing = await db.categories.find_one({"name": data.name, "id": {"$ne": category_id}})
        if existing:
            raise HTTPException(status_code=400, detail="Ilyen nevű kategória már létezik")
        update_data["name"] = data.name
        # Update all workers with old category name
        await db.workers.update_many(
            {"category": old_name},
            {"$set": {"category": data.name}}
        )
    
    if data.color is not None:
        update_data["color"] = data.color
    
    if data.order is not None:
        update_data["order"] = data.order
    
    if update_data:
        await db.categories.update_one({"id": category_id}, {"$set": update_data})
    
    updated = await db.categories.find_one({"id": category_id}, {"_id": 0})
    worker_count = await db.workers.count_documents({"category": updated["name"]})
    return CategoryResponse(**updated, worker_count=worker_count)

@api_router.delete("/categories/{category_id}")
async def delete_category(category_id: str, user: dict = Depends(require_admin)):
    """Kategória törlése - csak admin"""
    # Ellenőrizzük, hogy nem használja-e dolgozó
    workers_using = await db.workers.count_documents({"category": {"$exists": True}})
    category = await db.categories.find_one({"id": category_id}, {"_id": 0})
    if category:
        workers_with_cat = await db.workers.count_documents({"category": category["name"]})
        if workers_with_cat > 0:
            raise HTTPException(
                status_code=400, 
                detail=f"Nem törölhető: {workers_with_cat} dolgozó használja ezt a kategóriát"
            )
    
    result = await db.categories.delete_one({"id": category_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Kategória nem található")
    return {"message": "Kategória törölve"}

# ==================== USERS (Admin) ====================

@api_router.get("/users", response_model=List[UserResponse])
async def get_users(user: dict = Depends(require_admin)):
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(100)
    return [UserResponse(
        id=u["id"],
        email=u["email"],
        name=u.get("name", ""),
        role=u["role"],
        created_at=u.get("created_at", "")
    ) for u in users]

@api_router.get("/users/stats")
async def get_user_stats(user: dict = Depends(require_admin)):
    """Toborzónként hány dolgozót vitt fel"""
    pipeline = [
        {"$group": {"_id": "$owner_id", "count": {"$sum": 1}}},
    ]
    stats = await db.workers.aggregate(pipeline).to_list(100)
    
    result = []
    for s in stats:
        owner = await db.users.find_one({"id": s["_id"]}, {"_id": 0, "password": 0})
        if owner:
            result.append({
                "user_id": s["_id"],
                "user_name": owner.get("name", owner["email"]),
                "user_email": owner["email"],
                "worker_count": s["count"]
            })
    return result

# ==================== WORKERS ====================

@api_router.get("/workers", response_model=List[WorkerResponse])
async def get_workers(
    search: Optional[str] = None,
    category: Optional[str] = None,
    worker_type_id: Optional[str] = None,
    tag_id: Optional[str] = None,
    owner_id: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    query = {}
    
    # Toborzó csak saját dolgozóit látja
    if user["role"] != "admin":
        query["owner_id"] = user["id"]
    elif owner_id:
        query["owner_id"] = owner_id
    
    if category:
        query["category"] = category
    if worker_type_id:
        query["worker_type_id"] = worker_type_id
    if tag_id:
        query["tag_ids"] = tag_id
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}},
            {"address": {"$regex": search, "$options": "i"}},
            {"experience": {"$regex": search, "$options": "i"}},
            {"position": {"$regex": search, "$options": "i"}}
        ]
    
    workers = await db.workers.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # Enrich with type names, tags, project statuses
    result = []
    for w in workers:
        # Get type name
        type_doc = await db.worker_types.find_one({"id": w.get("worker_type_id")}, {"_id": 0})
        w["worker_type_name"] = type_doc["name"] if type_doc else ""
        
        # Position is now free text
        w["position"] = w.get("position", "")
        w["position_experience"] = w.get("position_experience", "")
        
        # Global status
        w["global_status"] = w.get("global_status", "Feldolgozatlan")
        
        # Get tags
        tag_ids = w.get("tag_ids", [])
        tags = []
        for tid in tag_ids:
            tag = await db.tags.find_one({"id": tid}, {"_id": 0})
            if tag:
                tags.append(tag)
        w["tags"] = tags
        
        # Get project statuses
        project_workers = await db.project_workers.find(
            {"worker_id": w["id"]}, {"_id": 0}
        ).sort("updated_at", -1).to_list(100)
        
        project_statuses = []
        for pw in project_workers:
            project = await db.projects.find_one({"id": pw["project_id"]}, {"_id": 0})
            status = await db.statuses.find_one({"id": pw.get("status_id")}, {"_id": 0})
            if project:
                project_statuses.append({
                    "project_id": project["id"],
                    "project_name": project["name"],
                    "project_date": project.get("date", ""),
                    "status_id": pw.get("status_id", ""),
                    "status_name": status["name"] if status else "Hozzárendelve",
                    "notes": pw.get("notes", ""),
                    "updated_at": pw.get("updated_at", "")
                })
        w["project_statuses"] = project_statuses
        
        # Get owner name
        owner = await db.users.find_one({"id": w.get("owner_id")}, {"_id": 0, "password": 0})
        w["owner_name"] = owner.get("name", owner["email"]) if owner else ""
        
        result.append(WorkerResponse(**w))
    
    return result

@api_router.get("/workers/{worker_id}", response_model=WorkerResponse)
async def get_worker(worker_id: str, user: dict = Depends(get_current_user)):
    query = {"id": worker_id}
    if user["role"] != "admin":
        query["owner_id"] = user["id"]
    
    w = await db.workers.find_one(query, {"_id": 0})
    if not w:
        raise HTTPException(status_code=404, detail="Dolgozó nem található")
    
    # Enrich
    type_doc = await db.worker_types.find_one({"id": w.get("worker_type_id")}, {"_id": 0})
    w["worker_type_name"] = type_doc["name"] if type_doc else ""
    
    # Position is now free text
    w["position"] = w.get("position", "")
    w["position_experience"] = w.get("position_experience", "")
    
    # Global status
    w["global_status"] = w.get("global_status", "Feldolgozatlan")
    
    tag_ids = w.get("tag_ids", [])
    tags = []
    for tid in tag_ids:
        tag = await db.tags.find_one({"id": tid}, {"_id": 0})
        if tag:
            tags.append(tag)
    w["tags"] = tags
    
    project_workers = await db.project_workers.find(
        {"worker_id": w["id"]}, {"_id": 0}
    ).sort("updated_at", -1).to_list(100)
    
    project_statuses = []
    for pw in project_workers:
        project = await db.projects.find_one({"id": pw["project_id"]}, {"_id": 0})
        status = await db.statuses.find_one({"id": pw.get("status_id")}, {"_id": 0})
        if project:
            project_statuses.append({
                "project_id": project["id"],
                "project_name": project["name"],
                "project_date": project.get("date", ""),
                "status_id": pw.get("status_id", ""),
                "status_name": status["name"] if status else "Hozzárendelve",
                "notes": pw.get("notes", ""),
                "updated_at": pw.get("updated_at", "")
            })
    w["project_statuses"] = project_statuses
    
    owner = await db.users.find_one({"id": w.get("owner_id")}, {"_id": 0, "password": 0})
    w["owner_name"] = owner.get("name", owner["email"]) if owner else ""
    
    return WorkerResponse(**w)

@api_router.post("/workers", response_model=WorkerResponse)
async def create_worker(data: WorkerCreate, user: dict = Depends(get_current_user)):
    if len(data.name) < 2:
        raise HTTPException(status_code=400, detail="A név minimum 2 karakter legyen")
    
    worker_doc = {
        "id": str(uuid.uuid4()),
        "name": data.name,
        "phone": data.phone,
        "worker_type_id": data.worker_type_id,
        "position": data.position or "",
        "position_experience": data.position_experience or "",
        "category": data.category,
        "address": data.address or "",
        "email": data.email or "",
        "experience": data.experience or "",
        "notes": data.notes or "",
        "global_status": data.global_status or "Feldolgozatlan",
        "tag_ids": [],
        "owner_id": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.workers.insert_one(worker_doc)
    
    # Ha project_id meg van adva, hozzáadjuk a várólistához
    if data.project_id:
        project = await db.projects.find_one({"id": data.project_id}, {"_id": 0})
        if project:
            # Ellenőrizzük, hogy a user hozzáfér-e a projekthez
            if user["role"] != "admin" and user["id"] not in project.get("recruiter_ids", []):
                raise HTTPException(status_code=403, detail="Nincs jogosultságod ehhez a projekthez")
            
            waitlist_doc = {
                "id": str(uuid.uuid4()),
                "project_id": data.project_id,
                "worker_id": worker_doc["id"],
                "start_date": data.start_date or "",
                "notes": "",
                "added_at": datetime.now(timezone.utc).isoformat(),
                "added_by": user["id"]
            }
            await db.project_waitlist.insert_one(waitlist_doc)
    
    worker_doc["worker_type_name"] = ""
    worker_doc["tags"] = []
    worker_doc["project_statuses"] = []
    worker_doc["owner_name"] = user.get("name", user["email"])
    
    return WorkerResponse(**worker_doc)

@api_router.put("/workers/{worker_id}", response_model=WorkerResponse)
async def update_worker(worker_id: str, data: WorkerUpdate, user: dict = Depends(get_current_user)):
    query = {"id": worker_id}
    if user["role"] != "admin":
        query["owner_id"] = user["id"]
    
    worker = await db.workers.find_one(query, {"_id": 0})
    if not worker:
        raise HTTPException(status_code=404, detail="Dolgozó nem található")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if update_data:
        await db.workers.update_one({"id": worker_id}, {"$set": update_data})
    
    return await get_worker(worker_id, user)

@api_router.delete("/workers/{worker_id}")
async def delete_worker(worker_id: str, user: dict = Depends(require_admin)):
    """Csak admin törölhet"""
    result = await db.workers.delete_one({"id": worker_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Dolgozó nem található")
    
    # Töröljük a projekt kapcsolatokat is
    await db.project_workers.delete_many({"worker_id": worker_id})
    
    return {"message": "Dolgozó törölve"}

@api_router.post("/workers/{worker_id}/tags/{tag_id}")
async def add_tag_to_worker(worker_id: str, tag_id: str, user: dict = Depends(get_current_user)):
    query = {"id": worker_id}
    if user["role"] != "admin":
        query["owner_id"] = user["id"]
    
    worker = await db.workers.find_one(query)
    if not worker:
        raise HTTPException(status_code=404, detail="Dolgozó nem található")
    
    await db.workers.update_one(
        {"id": worker_id},
        {"$addToSet": {"tag_ids": tag_id}}
    )
    return {"message": "Jellemző hozzáadva"}

@api_router.delete("/workers/{worker_id}/tags/{tag_id}")
async def remove_tag_from_worker(worker_id: str, tag_id: str, user: dict = Depends(get_current_user)):
    query = {"id": worker_id}
    if user["role"] != "admin":
        query["owner_id"] = user["id"]
    
    worker = await db.workers.find_one(query)
    if not worker:
        raise HTTPException(status_code=404, detail="Dolgozó nem található")
    
    await db.workers.update_one(
        {"id": worker_id},
        {"$pull": {"tag_ids": tag_id}}
    )
    return {"message": "Jellemző eltávolítva"}

# ==================== PROJECTS ====================

@api_router.get("/projects", response_model=List[ProjectResponse])
async def get_projects(user: dict = Depends(get_current_user)):
    """Toborzó csak azokat a projekteket látja, ahol ő hozta létre VAGY hozzá van rendelve"""
    projects = await db.projects.find({}, {"_id": 0}).sort("date", -1).to_list(1000)
    
    result = []
    for p in projects:
        count = await db.project_workers.count_documents({"project_id": p["id"]})
        position_count = await db.project_positions.count_documents({"project_id": p["id"]})
        trial_count = await db.trials.count_documents({"project_id": p["id"]})
        
        # Calculate total headcount from positions
        positions = await db.project_positions.find({"project_id": p["id"]}, {"_id": 0}).to_list(100)
        total_headcount = sum(pos.get("headcount", 0) for pos in positions)
        
        recruiter_ids = p.get("recruiter_ids", [])
        owner_id = p.get("owner_id", "")
        
        # Ha toborzó, csak azokat mutassa ahol ő hozta létre VAGY hozzá van rendelve
        if user["role"] != "admin":
            if owner_id != user["id"] and user["id"] not in recruiter_ids:
                continue
        
        # Get recruiter names
        recruiters = []
        for rid in recruiter_ids:
            r = await db.users.find_one({"id": rid}, {"_id": 0, "password": 0})
            if r:
                recruiters.append({"id": r["id"], "name": r.get("name", r["email"]), "email": r["email"]})
        
        # Get owner name
        owner_name = ""
        if owner_id:
            owner = await db.users.find_one({"id": owner_id}, {"_id": 0, "password": 0})
            if owner:
                owner_name = owner.get("name", owner["email"])
        
        result.append(ProjectResponse(
            id=p["id"],
            name=p["name"],
            client_name=p.get("client_name", ""),
            date=p["date"],
            location=p.get("location", ""),
            training_location=p.get("training_location", ""),
            notes=p.get("notes", ""),
            is_closed=p.get("is_closed", False),
            worker_count=count,
            position_count=position_count,
            total_headcount=total_headcount,
            trial_count=trial_count,
            recruiter_ids=recruiter_ids,
            recruiters=recruiters,
            owner_id=owner_id,
            owner_name=owner_name,
            created_at=p.get("created_at", "")
        ))
    
    return result

@api_router.get("/projects/{project_id}")
async def get_project(project_id: str, user: dict = Depends(get_current_user)):
    p = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    # Ellenőrizzük jogosultságot
    recruiter_ids = p.get("recruiter_ids", [])
    owner_id = p.get("owner_id", "")
    if user["role"] != "admin" and owner_id != user["id"] and user["id"] not in recruiter_ids:
        raise HTTPException(status_code=403, detail="Nincs hozzáférésed ehhez a projekthez")
    
    # Get workers
    pw_list = await db.project_workers.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    
    workers = []
    for pw in pw_list:
        w = await db.workers.find_one({"id": pw["worker_id"]}, {"_id": 0})
        if w:
            # Toborzó csak saját dolgozóit látja a projektben
            if user["role"] != "admin" and w.get("owner_id") != user["id"]:
                continue
            status = await db.statuses.find_one({"id": pw.get("status_id")}, {"_id": 0})
            type_doc = await db.worker_types.find_one({"id": w.get("worker_type_id")}, {"_id": 0})
            owner = await db.users.find_one({"id": w.get("owner_id")}, {"_id": 0, "password": 0})
            workers.append({
                "id": w["id"],
                "name": w["name"],
                "phone": w["phone"],
                "category": w["category"],
                "global_status": w.get("global_status", "Feldolgozatlan"),
                "worker_type_name": type_doc["name"] if type_doc else "",
                "status_id": pw.get("status_id", ""),
                "status_name": status["name"] if status else "Hozzárendelve",
                "notes": pw.get("notes", ""),
                "added_by": owner.get("name", owner["email"]) if owner else "",
                "added_at": pw.get("created_at", "")
            })
    
    total_count = await db.project_workers.count_documents({"project_id": project_id})
    
    # Get positions
    positions = await db.project_positions.find({"project_id": project_id}, {"_id": 0}).to_list(100)
    total_headcount = sum(pos.get("headcount", 0) for pos in positions)
    
    # Get trials with positions
    trials = await db.trials.find({"project_id": project_id}, {"_id": 0}).sort("date", 1).to_list(100)
    for trial in trials:
        trial_workers = await db.trial_workers.find({"trial_id": trial["id"]}, {"_id": 0}).to_list(100)
        trial["worker_count"] = len(trial_workers)
        
        # Get trial positions
        trial_positions = await db.trial_positions.find({"trial_id": trial["id"]}, {"_id": 0}).to_list(50)
        trial_pos_response = []
        for tp in trial_positions:
            assigned = await db.trial_workers.count_documents({
                "trial_id": trial["id"],
                "trial_position_id": tp["id"]
            })
            trial_pos_response.append({
                **tp,
                "assigned_count": assigned
            })
        trial["positions"] = trial_pos_response
        
        # Get worker details
        workers_list = []
        for tw in trial_workers:
            w = await db.workers.find_one({"id": tw["worker_id"]}, {"_id": 0})
            if w:
                trial_pos = await db.trial_positions.find_one({"id": tw.get("trial_position_id")}, {"_id": 0}) if tw.get("trial_position_id") else None
                workers_list.append({
                    "id": w["id"],
                    "name": w["name"],
                    "phone": w.get("phone", ""),
                    "trial_position_id": tw.get("trial_position_id", ""),
                    "position_name": trial_pos["position_name"] if trial_pos else ""
                })
        trial["workers"] = workers_list
    
    # Get recruiter names
    recruiters = []
    for rid in recruiter_ids:
        r = await db.users.find_one({"id": rid}, {"_id": 0, "password": 0})
        if r:
            recruiters.append({"id": r["id"], "name": r.get("name", r["email"]), "email": r["email"]})
    
    # Get owner name
    owner_name = ""
    if owner_id:
        owner = await db.users.find_one({"id": owner_id}, {"_id": 0, "password": 0})
        if owner:
            owner_name = owner.get("name", owner["email"])
    
    return {
        "id": p["id"],
        "name": p["name"],
        "client_name": p.get("client_name", ""),
        "date": p["date"],
        "location": p.get("location", ""),
        "training_location": p.get("training_location", ""),
        "notes": p.get("notes", ""),
        "is_closed": p.get("is_closed", False),
        "worker_count": total_count,
        "total_headcount": total_headcount,
        "recruiter_ids": recruiter_ids,
        "recruiters": recruiters,
        "owner_id": owner_id,
        "owner_name": owner_name,
        "workers": workers,
        "positions": positions,
        "trials": trials,
        "created_at": p.get("created_at", "")
    }

@api_router.post("/projects", response_model=ProjectResponse)
async def create_project(data: ProjectCreate, user: dict = Depends(require_admin)):
    """Csak admin hozhat létre projektet"""
    project_doc = {
        "id": str(uuid.uuid4()),
        "name": data.name,
        "client_name": data.client_name or "",
        "date": data.date,
        "location": data.location or "",
        "training_location": data.training_location or "",
        "notes": data.notes or "",
        "recruiter_ids": data.recruiter_ids,
        "is_closed": False,
        "owner_id": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.projects.insert_one(project_doc)
    
    owner_name = user.get("name", user["email"])
    return ProjectResponse(**project_doc, worker_count=0, position_count=0, total_headcount=0, trial_count=0, recruiters=[], owner_name=owner_name)

@api_router.put("/projects/{project_id}", response_model=ProjectResponse)
async def update_project(project_id: str, data: ProjectUpdate, user: dict = Depends(get_current_user)):
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if update_data:
        await db.projects.update_one({"id": project_id}, {"$set": update_data})
    
    updated = await db.projects.find_one({"id": project_id}, {"_id": 0})
    count = await db.project_workers.count_documents({"project_id": project_id})
    position_count = await db.project_positions.count_documents({"project_id": project_id})
    trial_count = await db.trials.count_documents({"project_id": project_id})
    
    # Calculate total headcount
    positions = await db.project_positions.find({"project_id": project_id}, {"_id": 0}).to_list(100)
    total_headcount = sum(pos.get("headcount", 0) for pos in positions)
    
    # Get recruiters
    recruiters = []
    for rid in updated.get("recruiter_ids", []):
        r = await db.users.find_one({"id": rid}, {"_id": 0, "password": 0})
        if r:
            recruiters.append({"id": r["id"], "name": r.get("name", r["email"]), "email": r["email"]})
    
    # Get owner name
    owner_name = ""
    owner_id = updated.get("owner_id", "")
    if owner_id:
        owner = await db.users.find_one({"id": owner_id}, {"_id": 0, "password": 0})
        if owner:
            owner_name = owner.get("name", owner["email"])
    
    return ProjectResponse(**updated, worker_count=count, position_count=position_count, total_headcount=total_headcount, trial_count=trial_count, recruiters=recruiters, owner_name=owner_name)

@api_router.post("/projects/{project_id}/recruiters")
async def add_recruiter_to_project(project_id: str, data: ProjectRecruiterAdd, user: dict = Depends(require_admin)):
    """Admin hozzárendel egy toborzót a projekthez"""
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    target_user = await db.users.find_one({"id": data.user_id}, {"_id": 0})
    if not target_user:
        raise HTTPException(status_code=404, detail="Felhasználó nem található")
    
    # Check if already assigned
    if data.user_id in project.get("recruiter_ids", []):
        return {"message": "Toborzó már hozzá van rendelve"}
    
    await db.projects.update_one(
        {"id": project_id},
        {"$addToSet": {"recruiter_ids": data.user_id}}
    )
    
    # Create notification for the recruiter
    await create_notification(
        user_id=data.user_id,
        notification_type="project_assigned",
        title="Új projekt hozzárendelés",
        message=f"Hozzárendeltek a(z) '{project['name']}' projekthez",
        link=f"/projects/{project_id}"
    )
    
    return {"message": "Toborzó hozzárendelve a projekthez"}

@api_router.delete("/projects/{project_id}/recruiters/{user_id}")
async def remove_recruiter_from_project(project_id: str, user_id: str, user: dict = Depends(require_admin)):
    """Admin eltávolít egy toborzót a projektből"""
    result = await db.projects.update_one(
        {"id": project_id},
        {"$pull": {"recruiter_ids": user_id}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    return {"message": "Toborzó eltávolítva a projektről"}

@api_router.delete("/projects/{project_id}")
async def delete_project(project_id: str, user: dict = Depends(require_admin)):
    """Csak admin törölhet projektet"""
    result = await db.projects.delete_one({"id": project_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    await db.project_workers.delete_many({"project_id": project_id})
    await db.project_positions.delete_many({"project_id": project_id})
    await db.trials.delete_many({"project_id": project_id})
    return {"message": "Projekt törölve"}

# ==================== PROJECT POSITIONS ====================

@api_router.get("/projects/{project_id}/positions", response_model=List[ProjectPositionResponse])
async def get_project_positions(project_id: str, user: dict = Depends(get_current_user)):
    """Projekt pozícióinak lekérése"""
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    positions = await db.project_positions.find({"project_id": project_id}, {"_id": 0}).to_list(100)
    
    result = []
    for pos in positions:
        # Count assigned workers for this position
        assigned_workers = await db.project_workers.count_documents({
            "project_id": project_id,
            "position_id": pos["id"]
        })
        result.append(ProjectPositionResponse(
            **pos,
            assigned_workers=assigned_workers
        ))
    
    return result

@api_router.post("/projects/{project_id}/positions", response_model=ProjectPositionResponse)
async def create_project_position(project_id: str, data: ProjectPositionCreate, user: dict = Depends(require_admin)):
    """Pozíció létrehozása projekthez - csak admin"""
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    position_doc = {
        "id": str(uuid.uuid4()),
        "project_id": project_id,
        "name": data.name,
        "headcount": data.headcount,
        "shift_schedule": data.shift_schedule or "",
        "experience_required": data.experience_required or "",
        "qualifications": data.qualifications or "",
        "physical_requirements": data.physical_requirements or "",
        "notes": data.notes or "",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.project_positions.insert_one(position_doc)
    
    return ProjectPositionResponse(**position_doc, assigned_workers=0)

@api_router.put("/projects/{project_id}/positions/{position_id}", response_model=ProjectPositionResponse)
async def update_project_position(project_id: str, position_id: str, data: ProjectPositionUpdate, user: dict = Depends(require_admin)):
    """Pozíció szerkesztése - csak admin"""
    position = await db.project_positions.find_one({"id": position_id, "project_id": project_id})
    if not position:
        raise HTTPException(status_code=404, detail="Pozíció nem található")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if update_data:
        await db.project_positions.update_one({"id": position_id}, {"$set": update_data})
    
    updated = await db.project_positions.find_one({"id": position_id}, {"_id": 0})
    assigned_workers = await db.project_workers.count_documents({
        "project_id": project_id,
        "position_id": position_id
    })
    
    return ProjectPositionResponse(**updated, assigned_workers=assigned_workers)

@api_router.delete("/projects/{project_id}/positions/{position_id}")
async def delete_project_position(project_id: str, position_id: str, user: dict = Depends(require_admin)):
    """Pozíció törlése - csak admin"""
    result = await db.project_positions.delete_one({"id": position_id, "project_id": project_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Pozíció nem található")
    
    # Remove position_id from project_workers
    await db.project_workers.update_many(
        {"project_id": project_id, "position_id": position_id},
        {"$unset": {"position_id": ""}}
    )
    
    return {"message": "Pozíció törölve"}

# ==================== WAITLIST (VÁRÓLISTA) ====================

@api_router.get("/projects/{project_id}/waitlist", response_model=List[WaitlistWorkerResponse])
async def get_project_waitlist(project_id: str, user: dict = Depends(get_current_user)):
    """Get all workers in project waitlist"""
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    # Check permission
    if user["role"] != "admin" and user["id"] not in project.get("recruiter_ids", []):
        raise HTTPException(status_code=403, detail="Nincs jogosultságod ehhez a projekthez")
    
    waitlist_entries = await db.project_waitlist.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    
    result = []
    for entry in waitlist_entries:
        worker = await db.workers.find_one({"id": entry["worker_id"]}, {"_id": 0})
        if worker:
            added_by_user = await db.users.find_one({"id": entry["added_by"]}, {"_id": 0})
            result.append(WaitlistWorkerResponse(
                id=entry["id"],
                project_id=entry["project_id"],
                worker_id=entry["worker_id"],
                worker_name=worker["name"],
                worker_phone=worker["phone"],
                worker_email=worker.get("email", ""),
                start_date=entry.get("start_date", ""),
                notes=entry.get("notes", ""),
                added_at=entry["added_at"],
                added_by=entry["added_by"],
                added_by_name=added_by_user.get("name", added_by_user["email"]) if added_by_user else ""
            ))
    
    return result

@api_router.post("/projects/{project_id}/waitlist")
async def add_worker_to_waitlist(project_id: str, data: WaitlistWorkerAdd, user: dict = Depends(get_current_user)):
    """Add a worker to project waitlist"""
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    # Check permission
    if user["role"] != "admin" and user["id"] not in project.get("recruiter_ids", []):
        raise HTTPException(status_code=403, detail="Nincs jogosultságod ehhez a projekthez")
    
    # Check if worker exists
    worker = await db.workers.find_one({"id": data.worker_id}, {"_id": 0})
    if not worker:
        raise HTTPException(status_code=404, detail="Dolgozó nem található")
    
    # Check if worker already in waitlist
    existing = await db.project_waitlist.find_one({"project_id": project_id, "worker_id": data.worker_id})
    if existing:
        raise HTTPException(status_code=400, detail="A dolgozó már a várólistán van")
    
    waitlist_doc = {
        "id": str(uuid.uuid4()),
        "project_id": project_id,
        "worker_id": data.worker_id,
        "start_date": data.start_date or "",
        "notes": data.notes or "",
        "added_at": datetime.now(timezone.utc).isoformat(),
        "added_by": user["id"]
    }
    await db.project_waitlist.insert_one(waitlist_doc)
    
    return {"message": "Dolgozó hozzáadva a várólistához", "id": waitlist_doc["id"]}

@api_router.put("/projects/{project_id}/waitlist/{worker_id}")
async def update_waitlist_entry(project_id: str, worker_id: str, data: WaitlistWorkerUpdate, user: dict = Depends(get_current_user)):
    """Update waitlist entry"""
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    if user["role"] != "admin" and user["id"] not in project.get("recruiter_ids", []):
        raise HTTPException(status_code=403, detail="Nincs jogosultságod")
    
    entry = await db.project_waitlist.find_one({"project_id": project_id, "worker_id": worker_id})
    if not entry:
        raise HTTPException(status_code=404, detail="Dolgozó nincs a várólistán")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if update_data:
        await db.project_waitlist.update_one(
            {"project_id": project_id, "worker_id": worker_id},
            {"$set": update_data}
        )
    
    return {"message": "Várólista frissítve"}

@api_router.delete("/projects/{project_id}/waitlist/{worker_id}")
async def remove_worker_from_waitlist(project_id: str, worker_id: str, user: dict = Depends(get_current_user)):
    """Remove worker from project waitlist"""
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    if user["role"] != "admin" and user["id"] not in project.get("recruiter_ids", []):
        raise HTTPException(status_code=403, detail="Nincs jogosultságod")
    
    result = await db.project_waitlist.delete_one({"project_id": project_id, "worker_id": worker_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Dolgozó nincs a várólistán")
    
    return {"message": "Dolgozó eltávolítva a várólistáról"}

# ==================== TRIALS (PRÓBÁK) ====================

@api_router.get("/projects/{project_id}/trials", response_model=List[TrialResponse])
async def get_project_trials(project_id: str, user: dict = Depends(get_current_user)):
    """Projekt próbáinak lekérése"""
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    trials = await db.trials.find({"project_id": project_id}, {"_id": 0}).sort("date", 1).to_list(100)
    
    result = []
    for trial in trials:
        # Get workers for this trial
        trial_workers_list = await db.trial_workers.find({"trial_id": trial["id"]}, {"_id": 0}).to_list(100)
        
        workers = []
        for tw in trial_workers_list:
            w = await db.workers.find_one({"id": tw["worker_id"]}, {"_id": 0})
            if w:
                # Toborzó csak saját dolgozóit látja
                if user["role"] != "admin" and w.get("owner_id") != user["id"]:
                    continue
                # Get trial position name if assigned
                trial_pos = await db.trial_positions.find_one({"id": tw.get("trial_position_id")}, {"_id": 0}) if tw.get("trial_position_id") else None
                workers.append({
                    "id": w["id"],
                    "name": w["name"],
                    "phone": w["phone"],
                    "trial_position_id": tw.get("trial_position_id", ""),
                    "position_name": trial_pos["position_name"] if trial_pos else "",
                    "added_at": tw.get("created_at", "")
                })
        
        # Get trial positions
        trial_positions = await db.trial_positions.find({"trial_id": trial["id"]}, {"_id": 0}).to_list(50)
        positions_response = []
        for tp in trial_positions:
            # Count assigned workers for this position
            assigned = await db.trial_workers.count_documents({
                "trial_id": trial["id"],
                "trial_position_id": tp["id"]
            })
            positions_response.append(TrialPositionResponse(
                **tp,
                assigned_count=assigned
            ))
        
        result.append(TrialResponse(
            **trial,
            worker_count=len(trial_workers_list),
            workers=workers,
            positions=positions_response
        ))
    
    return result

@api_router.post("/projects/{project_id}/trials", response_model=TrialResponse)
async def create_trial(project_id: str, data: TrialCreate, user: dict = Depends(require_admin)):
    """Próba létrehozása - csak admin"""
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    trial_doc = {
        "id": str(uuid.uuid4()),
        "project_id": project_id,
        "date": data.date,
        "time": data.time or "",
        "notes": data.notes or "",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.trials.insert_one(trial_doc)
    
    return TrialResponse(**trial_doc, worker_count=0, workers=[], positions=[])

@api_router.put("/projects/{project_id}/trials/{trial_id}", response_model=TrialResponse)
async def update_trial(project_id: str, trial_id: str, data: TrialUpdate, user: dict = Depends(require_admin)):
    """Próba szerkesztése - csak admin"""
    trial = await db.trials.find_one({"id": trial_id, "project_id": project_id})
    if not trial:
        raise HTTPException(status_code=404, detail="Próba nem található")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if update_data:
        await db.trials.update_one({"id": trial_id}, {"$set": update_data})
    
    updated = await db.trials.find_one({"id": trial_id}, {"_id": 0})
    worker_count = await db.trial_workers.count_documents({"trial_id": trial_id})
    
    return TrialResponse(**updated, worker_count=worker_count, workers=[])

@api_router.delete("/projects/{project_id}/trials/{trial_id}")
async def delete_trial(project_id: str, trial_id: str, user: dict = Depends(require_admin)):
    """Próba törlése - csak admin"""
    result = await db.trials.delete_one({"id": trial_id, "project_id": project_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Próba nem található")
    
    # Delete trial workers and positions
    await db.trial_workers.delete_many({"trial_id": trial_id})
    await db.trial_positions.delete_many({"trial_id": trial_id})
    
    return {"message": "Próba törölve"}

# ==================== TRIAL POSITIONS ====================

@api_router.get("/projects/{project_id}/trials/{trial_id}/positions")
async def get_trial_positions(project_id: str, trial_id: str, user: dict = Depends(get_current_user)):
    """Próba pozícióinak lekérése"""
    trial = await db.trials.find_one({"id": trial_id, "project_id": project_id})
    if not trial:
        raise HTTPException(status_code=404, detail="Próba nem található")
    
    positions = await db.trial_positions.find({"trial_id": trial_id}, {"_id": 0}).to_list(50)
    result = []
    for tp in positions:
        assigned = await db.trial_workers.count_documents({
            "trial_id": trial_id,
            "trial_position_id": tp["id"]
        })
        result.append(TrialPositionResponse(**tp, assigned_count=assigned))
    return result

@api_router.post("/projects/{project_id}/trials/{trial_id}/positions", response_model=TrialPositionResponse)
async def add_trial_position(project_id: str, trial_id: str, data: TrialPositionCreate, user: dict = Depends(get_current_user)):
    """Pozíció hozzáadása próbához"""
    trial = await db.trials.find_one({"id": trial_id, "project_id": project_id})
    if not trial:
        raise HTTPException(status_code=404, detail="Próba nem található")
    
    # Check if position name already exists in trial
    existing = await db.trial_positions.find_one({
        "trial_id": trial_id,
        "position_name": data.position_name
    })
    if existing:
        raise HTTPException(status_code=400, detail="Ez a pozíció már hozzá van adva ehhez a próbához")
    
    # If add_to_project is True, add to project positions too
    if data.add_to_project:
        existing_project_pos = await db.project_positions.find_one({
            "project_id": project_id,
            "name": data.position_name
        })
        if not existing_project_pos:
            project_pos_doc = {
                "id": str(uuid.uuid4()),
                "project_id": project_id,
                "name": data.position_name,
                "headcount": data.headcount,
                "shift_schedule": "",
                "experience_required": data.requirements,
                "qualifications": "",
                "physical_requirements": "",
                "notes": "",
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.project_positions.insert_one(project_pos_doc)
            data.position_id = project_pos_doc["id"]
    
    trial_pos_doc = {
        "id": str(uuid.uuid4()),
        "trial_id": trial_id,
        "position_id": data.position_id or "",
        "position_name": data.position_name,
        "headcount": data.headcount,
        "hourly_rate": data.hourly_rate or "",
        "accommodation": data.accommodation,
        "requirements": data.requirements,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.trial_positions.insert_one(trial_pos_doc)
    
    return TrialPositionResponse(**trial_pos_doc, assigned_count=0)

@api_router.put("/projects/{project_id}/trials/{trial_id}/positions/{position_id}")
async def update_trial_position(project_id: str, trial_id: str, position_id: str, data: TrialPositionCreate, user: dict = Depends(get_current_user)):
    """Próba pozíció szerkesztése"""
    trial_pos = await db.trial_positions.find_one({"id": position_id, "trial_id": trial_id})
    if not trial_pos:
        raise HTTPException(status_code=404, detail="Pozíció nem található")
    
    update_data = {
        "position_name": data.position_name,
        "headcount": data.headcount,
        "hourly_rate": data.hourly_rate or "",
        "accommodation": data.accommodation,
        "requirements": data.requirements
    }
    await db.trial_positions.update_one({"id": position_id}, {"$set": update_data})
    
    updated = await db.trial_positions.find_one({"id": position_id}, {"_id": 0})
    assigned = await db.trial_workers.count_documents({"trial_id": trial_id, "trial_position_id": position_id})
    return TrialPositionResponse(**updated, assigned_count=assigned)

@api_router.delete("/projects/{project_id}/trials/{trial_id}/positions/{position_id}")
async def delete_trial_position(project_id: str, trial_id: str, position_id: str, user: dict = Depends(get_current_user)):
    """Próba pozíció törlése"""
    result = await db.trial_positions.delete_one({"id": position_id, "trial_id": trial_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Pozíció nem található")
    
    # Remove position assignment from workers
    await db.trial_workers.update_many(
        {"trial_id": trial_id, "trial_position_id": position_id},
        {"$set": {"trial_position_id": ""}}
    )
    
    return {"message": "Pozíció törölve"}

# ==================== TRIAL WORKERS ====================

@api_router.post("/projects/{project_id}/trials/{trial_id}/workers")
async def add_worker_to_trial(project_id: str, trial_id: str, data: TrialWorkerAdd, user: dict = Depends(get_current_user)):
    """Dolgozó hozzáadása próbához"""
    trial = await db.trials.find_one({"id": trial_id, "project_id": project_id}, {"_id": 0})
    if not trial:
        raise HTTPException(status_code=404, detail="Próba nem található")
    
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    
    worker = await db.workers.find_one({"id": data.worker_id}, {"_id": 0})
    if not worker:
        raise HTTPException(status_code=404, detail="Dolgozó nem található")
    
    existing = await db.trial_workers.find_one({
        "trial_id": trial_id,
        "worker_id": data.worker_id
    })
    if existing:
        raise HTTPException(status_code=400, detail="Dolgozó már hozzá van rendelve ehhez a próbához")
    
    tw_doc = {
        "id": str(uuid.uuid4()),
        "trial_id": trial_id,
        "worker_id": data.worker_id,
        "trial_position_id": data.position_id or "",  # This is now trial_position_id
        "added_by": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.trial_workers.insert_one(tw_doc)
    
    # Send notification to worker's owner (recruiter) if different from current user
    if worker.get("owner_id") and worker["owner_id"] != user["id"]:
        await create_notification(
            user_id=worker["owner_id"],
            notification_type="trial_assigned",
            title="Dolgozó próbára beosztva",
            message=f"A(z) '{worker['name']}' dolgozódat beosztották a '{project['name'] if project else 'Projekt'}' próbájára ({trial['date']})",
            link=f"/projects/{project_id}"
        )
    
    return {"message": "Dolgozó hozzáadva a próbához"}

@api_router.put("/projects/{project_id}/trials/{trial_id}/workers/{worker_id}/position")
async def assign_worker_to_trial_position(project_id: str, trial_id: str, worker_id: str, trial_position_id: str = "", user: dict = Depends(get_current_user)):
    """Dolgozó pozícióhoz rendelése a próbán belül"""
    result = await db.trial_workers.update_one(
        {"trial_id": trial_id, "worker_id": worker_id},
        {"$set": {"trial_position_id": trial_position_id}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Dolgozó nem található ezen a próbán")
    return {"message": "Pozíció frissítve"}

@api_router.delete("/projects/{project_id}/trials/{trial_id}/workers/{worker_id}")
async def remove_worker_from_trial(project_id: str, trial_id: str, worker_id: str, user: dict = Depends(get_current_user)):
    """Dolgozó eltávolítása próbáról"""
    result = await db.trial_workers.delete_one({
        "trial_id": trial_id,
        "worker_id": worker_id
    })
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Kapcsolat nem található")
    return {"message": "Dolgozó eltávolítva a próbáról"}

@api_router.post("/projects/{project_id}/workers")
async def add_worker_to_project(project_id: str, data: ProjectWorkerAdd, user: dict = Depends(get_current_user)):
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Projekt nem található")
    
    worker = await db.workers.find_one({"id": data.worker_id})
    if not worker:
        raise HTTPException(status_code=404, detail="Dolgozó nem található")
    
    existing = await db.project_workers.find_one({
        "project_id": project_id,
        "worker_id": data.worker_id
    })
    if existing:
        raise HTTPException(status_code=400, detail="Dolgozó már hozzá van rendelve")
    
    pw_doc = {
        "id": str(uuid.uuid4()),
        "project_id": project_id,
        "worker_id": data.worker_id,
        "status_id": data.status_id or "",
        "added_by": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.project_workers.insert_one(pw_doc)
    return {"message": "Dolgozó hozzáadva a projekthez"}

@api_router.delete("/projects/{project_id}/workers/{worker_id}")
async def remove_worker_from_project(project_id: str, worker_id: str, user: dict = Depends(get_current_user)):
    result = await db.project_workers.delete_one({
        "project_id": project_id,
        "worker_id": worker_id
    })
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Kapcsolat nem található")
    return {"message": "Dolgozó eltávolítva a projektről"}

@api_router.put("/projects/{project_id}/workers/{worker_id}/status")
async def update_worker_status_in_project(
    project_id: str, 
    worker_id: str, 
    data: ProjectWorkerStatusUpdate,
    user: dict = Depends(get_current_user)
):
    update_fields = {
        "status_id": data.status_id,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    if data.notes is not None:
        update_fields["notes"] = data.notes
    
    result = await db.project_workers.update_one(
        {"project_id": project_id, "worker_id": worker_id},
        {"$set": update_fields}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Kapcsolat nem található")
    return {"message": "Státusz frissítve"}

# ==================== EXCEL EXPORT ====================

async def generate_excel_for_user(user_id: str, user_name: str):
    """Generate Excel file for a specific recruiter with workers grouped by category"""
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Get categories from database
    db_categories = await db.categories.find({}, {"_id": 0}).to_list(100)
    categories = [c["name"] for c in db_categories] if db_categories else [
        "Felvitt dolgozók", "Hideg jelentkező", "Űrlapon jelentkezett", 
        "Állásra jelentkezett", "Ingázó", "Szállásos"
    ]
    
    # Remove default sheet
    wb.remove(wb.active)
    
    for cat in categories:
        workers = await db.workers.find(
            {"owner_id": user_id, "category": cat}, {"_id": 0}
        ).sort("name", 1).to_list(1000)
        
        if not workers:
            continue
            
        # Create sheet for category
        ws = wb.create_sheet(title=cat[:31])  # Excel max 31 chars
        
        # Headers
        headers = ["Név", "Telefon", "Email", "Lakcím", "Típus", "Tapasztalat", "Felvéve"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        for row, worker in enumerate(workers, 2):
            type_doc = await db.worker_types.find_one({"id": worker.get("worker_type_id")}, {"_id": 0})
            type_name = type_doc["name"] if type_doc else ""
            
            ws.cell(row=row, column=1, value=worker["name"]).border = border
            ws.cell(row=row, column=2, value=worker["phone"]).border = border
            ws.cell(row=row, column=3, value=worker.get("email", "")).border = border
            ws.cell(row=row, column=4, value=worker.get("address", "")).border = border
            ws.cell(row=row, column=5, value=type_name).border = border
            ws.cell(row=row, column=6, value=worker.get("experience", "")).border = border
            ws.cell(row=row, column=7, value=worker.get("created_at", "")[:10]).border = border
        
        # Auto-width columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    # If no sheets were created, add summary
    if not wb.sheetnames:
        ws = wb.create_sheet(title="Összefoglaló")
        ws.cell(row=1, column=1, value="Nincs dolgozó ebben a kategóriában")
    
    # Save file
    safe_name = "".join(c for c in user_name if c.isalnum() or c in " -_").strip() or "export"
    filename = f"{safe_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = EXPORTS_DIR / filename
    wb.save(filepath)
    
    return filepath, filename

@api_router.get("/export/workers")
async def export_workers_excel(user: dict = Depends(get_current_user)):
    """Export current user's workers to Excel"""
    user_name = user.get("name") or user["email"].split("@")[0]
    filepath, filename = await generate_excel_for_user(user["id"], user_name)
    
    return FileResponse(
        path=filepath,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@api_router.get("/export/workers/{user_id}")
async def export_user_workers_excel(user_id: str, admin: dict = Depends(require_admin)):
    """Admin can export any user's workers to Excel"""
    target_user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
    if not target_user:
        raise HTTPException(status_code=404, detail="Felhasználó nem található")
    
    user_name = target_user.get("name") or target_user["email"].split("@")[0]
    filepath, filename = await generate_excel_for_user(user_id, user_name)
    
    return FileResponse(
        path=filepath,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@api_router.get("/export/all")
async def export_all_workers_excel(admin: dict = Depends(require_admin)):
    """Admin exports all workers grouped by recruiter and category"""
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    recruiter_fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Get all users
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(100)
    
    # Remove default sheet
    wb.remove(wb.active)
    
    for u in users:
        workers = await db.workers.find({"owner_id": u["id"]}, {"_id": 0}).sort("category", 1).to_list(1000)
        
        if not workers:
            continue
        
        user_name = u.get("name") or u["email"].split("@")[0]
        sheet_name = user_name[:31]  # Excel max 31 chars
        
        # Handle duplicate sheet names
        if sheet_name in wb.sheetnames:
            sheet_name = f"{sheet_name[:28]}_{len(wb.sheetnames)}"
        
        ws = wb.create_sheet(title=sheet_name)
        
        # Headers
        headers = ["Név", "Telefon", "Email", "Kategória", "Típus", "Lakcím", "Felvéve"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Data
        for row, worker in enumerate(workers, 2):
            type_doc = await db.worker_types.find_one({"id": worker.get("worker_type_id")}, {"_id": 0})
            type_name = type_doc["name"] if type_doc else ""
            
            ws.cell(row=row, column=1, value=worker["name"]).border = border
            ws.cell(row=row, column=2, value=worker["phone"]).border = border
            ws.cell(row=row, column=3, value=worker.get("email", "")).border = border
            ws.cell(row=row, column=4, value=worker["category"]).border = border
            ws.cell(row=row, column=5, value=type_name).border = border
            ws.cell(row=row, column=6, value=worker.get("address", "")).border = border
            ws.cell(row=row, column=7, value=worker.get("created_at", "")[:10]).border = border
        
        # Auto-width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    if not wb.sheetnames:
        ws = wb.create_sheet(title="Összefoglaló")
        ws.cell(row=1, column=1, value="Nincs dolgozó a rendszerben")
    
    filename = f"osszes_dolgozo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = EXPORTS_DIR / filename
    wb.save(filepath)
    
    return FileResponse(
        path=filepath,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ==================== SEED DATA ====================

@api_router.post("/seed")
async def seed_data():
    """Initialize default data"""
    # Check if already seeded
    admin = await db.users.find_one({"email": "admin@dolgozocrm.hu"})
    if admin:
        return {"message": "Adatok már léteznek"}
    
    # Create admin user
    admin_doc = {
        "id": str(uuid.uuid4()),
        "email": "admin@dolgozocrm.hu",
        "password": hash_password("admin123"),
        "name": "Admin",
        "role": "admin",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(admin_doc)
    
    # Create test recruiter
    recruiter_doc = {
        "id": str(uuid.uuid4()),
        "email": "toborzo@dolgozocrm.hu",
        "password": hash_password("toborzo123"),
        "name": "Teszt Toborzó",
        "role": "user",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(recruiter_doc)
    
    # Worker types with positions
    type_positions = {
        "Betanított munkás": ["Csomagoló", "Komissiózó", "Összeszerelő", "Gyártósori munkás"],
        "Szakmunkás": ["Hegesztő", "Villanyszerelő", "Lakatos", "Esztergályos", "CNC gépkezelő", "Szerszámkészítő"],
        "Targoncás": ["Homlok targoncás", "Oldal targoncás", "Reach truck kezelő", "Magasraktári targoncás"],
        "Gépkezelő": ["Présgép kezelő", "Fröccsöntő gép kezelő", "Hajlítógép kezelő", "Varrógép kezelő"],
        "Raktáros": ["Áruátvevő", "Kiadó", "Leltáros", "Raktári adminisztrátor"],
        "Segédmunkás": ["Takarító", "Anyagmozgató", "Betanított segéd", "Kézi rakodó"]
    }
    
    for type_name, positions in type_positions.items():
        type_id = str(uuid.uuid4())
        await db.worker_types.insert_one({"id": type_id, "name": type_name})
        for pos in positions:
            await db.positions.insert_one({
                "id": str(uuid.uuid4()),
                "name": pos,
                "worker_type_id": type_id
            })
    
    # Statuses
    statuses = ["Jelentkezett", "Megerősítve", "Dolgozik", "Megfelelt", "Nem felelt meg", "Lemondta", "Nem jelent meg"]
    for s in statuses:
        await db.statuses.insert_one({"id": str(uuid.uuid4()), "name": s})
    
    # Tags
    tags = [
        {"name": "Megbízható", "color": "#22c55e"},
        {"name": "Tapasztalt", "color": "#3b82f6"},
        {"name": "Ajánlott", "color": "#f97316"},
        {"name": "Saját autó", "color": "#8b5cf6"},
        {"name": "Éjszakás", "color": "#6366f1"}
    ]
    for t in tags:
        await db.tags.insert_one({"id": str(uuid.uuid4()), **t})
    
    return {"message": "Seed adatok létrehozva", "admin_email": "admin@dolgozocrm.hu", "admin_password": "admin123"}

# ==================== EXCEL IMPORT ====================

class ExcelColumnMapping(BaseModel):
    name: Optional[int] = None  # Column index for name
    phone: Optional[int] = None
    email: Optional[int] = None
    address: Optional[int] = None
    position: Optional[int] = None
    experience: Optional[int] = None
    notes: Optional[int] = None

class ExcelImportSettings(BaseModel):
    column_mapping: Dict[str, Optional[int]]  # field_name -> column_index
    worker_type_id: str  # Required worker type
    category: str = "Felvitt dolgozók"  # Default category for all
    global_status: str = "Feldolgozatlan"
    start_row: int = 2  # Skip header row
    apply_same_to_all: bool = True  # Apply same category/type to all

@api_router.post("/workers/import/preview")
async def preview_excel_import(
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user)
):
    """Preview Excel file - returns first 10 rows and column headers"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Csak .xlsx vagy .xls fájl tölthető fel")
    
    try:
        contents = await file.read()
        wb = load_workbook(filename=io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
        
        # Get headers (first row)
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            headers.append(str(cell_value) if cell_value else f"Oszlop {col}")
        
        # Get preview rows (first 10 data rows)
        preview_rows = []
        for row in range(2, min(12, ws.max_row + 1)):
            row_data = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                row_data.append(str(cell_value) if cell_value else "")
            preview_rows.append(row_data)
        
        wb.close()
        
        return {
            "filename": file.filename,
            "total_rows": ws.max_row - 1,  # Exclude header
            "columns": headers,
            "preview_rows": preview_rows,
            "column_count": len(headers)
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Hiba a fájl olvasásakor: {str(e)}")

@api_router.post("/workers/import")
async def import_workers_from_excel(
    file: UploadFile = File(...),
    settings: str = Form(...),  # JSON string of ExcelImportSettings
    user: dict = Depends(get_current_user)
):
    """Import workers from Excel file with column mapping"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Csak .xlsx vagy .xls fájl tölthető fel")
    
    try:
        import_settings = json.loads(settings)
    except:
        raise HTTPException(status_code=400, detail="Érvénytelen beállítások")
    
    column_mapping = import_settings.get("column_mapping", {})
    worker_type_id = import_settings.get("worker_type_id", "")
    category = import_settings.get("category", "Felvitt dolgozók")
    global_status = import_settings.get("global_status", "Feldolgozatlan")
    start_row = import_settings.get("start_row", 2)
    
    # Validate worker_type_id
    worker_type = await db.worker_types.find_one({"id": worker_type_id})
    if not worker_type:
        raise HTTPException(status_code=400, detail="Érvénytelen dolgozó típus")
    
    # Name column is required
    name_col = column_mapping.get("name")
    if name_col is None:
        raise HTTPException(status_code=400, detail="Név oszlop megadása kötelező")
    
    try:
        contents = await file.read()
        wb = load_workbook(filename=io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
        
        imported_count = 0
        skipped_count = 0
        errors = []
        
        # Process rows (up to 1000)
        max_rows = min(ws.max_row, start_row + 1000)
        
        for row_idx in range(start_row, max_rows + 1):
            try:
                # Get name (required)
                name_value = ws.cell(row=row_idx, column=name_col + 1).value  # +1 because openpyxl is 1-indexed
                if not name_value or str(name_value).strip() == "":
                    skipped_count += 1
                    continue
                
                name = str(name_value).strip()
                if len(name) < 2:
                    skipped_count += 1
                    continue
                
                # Get optional fields
                phone_col = column_mapping.get("phone")
                phone = str(ws.cell(row=row_idx, column=phone_col + 1).value or "").strip() if phone_col is not None else ""
                
                email_col = column_mapping.get("email")
                email = str(ws.cell(row=row_idx, column=email_col + 1).value or "").strip() if email_col is not None else ""
                
                address_col = column_mapping.get("address")
                address = str(ws.cell(row=row_idx, column=address_col + 1).value or "").strip() if address_col is not None else ""
                
                position_col = column_mapping.get("position")
                position = str(ws.cell(row=row_idx, column=position_col + 1).value or "").strip() if position_col is not None else ""
                
                experience_col = column_mapping.get("experience")
                experience = str(ws.cell(row=row_idx, column=experience_col + 1).value or "").strip() if experience_col is not None else ""
                
                notes_col = column_mapping.get("notes")
                notes = str(ws.cell(row=row_idx, column=notes_col + 1).value or "").strip() if notes_col is not None else ""
                
                # Create worker document
                worker_doc = {
                    "id": str(uuid.uuid4()),
                    "name": name,
                    "phone": phone,
                    "worker_type_id": worker_type_id,
                    "position": position,
                    "position_experience": "",
                    "category": category,
                    "address": address,
                    "email": email,
                    "experience": experience,
                    "notes": notes,
                    "global_status": global_status,
                    "tag_ids": [],
                    "owner_id": user["id"],
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                
                await db.workers.insert_one(worker_doc)
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Sor {row_idx}: {str(e)}")
                skipped_count += 1
        
        wb.close()
        
        return {
            "message": f"{imported_count} dolgozó sikeresen importálva",
            "imported": imported_count,
            "skipped": skipped_count,
            "errors": errors[:10] if errors else []  # Return first 10 errors
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Hiba az importálás során: {str(e)}")

# ==================== FTP SYNC / BACKUP ====================

def generate_recruiter_excel(recruiter_id: str, recruiter_name: str, workers: list) -> bytes:
    """Generate Excel file with all worker data for a recruiter"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Dolgozók"
    
    # Headers
    headers = [
        "Teljes név", "Telefon", "Email", "Lakcím", "Pozíció", 
        "Tapasztalat", "Kategória", "Típus", "Globális státusz",
        "Projektek", "Megjegyzés", "Létrehozva"
    ]
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data
    for row_idx, w in enumerate(workers, 2):
        # Get project statuses as string
        project_info = ""
        if w.get("project_statuses"):
            project_info = "; ".join([
                f"{ps.get('project_name', '')}: {ps.get('status_name', '')}" 
                for ps in w.get("project_statuses", [])
            ])
        
        row_data = [
            w.get("name", ""),
            w.get("phone", ""),
            w.get("email", ""),
            w.get("address", ""),
            w.get("position", ""),
            w.get("experience", ""),
            w.get("category", ""),
            w.get("worker_type_name", ""),
            w.get("global_status", "Feldolgozatlan"),
            project_info,
            w.get("notes", ""),
            w.get("created_at", "")[:10] if w.get("created_at") else ""
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
    
    # Adjust column widths
    column_widths = [20, 15, 25, 30, 15, 20, 15, 15, 15, 40, 30, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

async def sync_to_ftp():
    """Sync all recruiter Excel files to FTP server"""
    if not FTP_HOST or not FTP_USER or not FTP_PASS:
        logger.warning("FTP credentials not configured, skipping sync")
        return {"status": "skipped", "reason": "FTP not configured"}
    
    try:
        # Get all users (recruiters)
        users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(100)
        
        # Connect to FTP
        ftp = ftplib.FTP(FTP_HOST)
        ftp.login(FTP_USER, FTP_PASS)
        
        # Try to create/change to backup folder
        try:
            ftp.cwd(FTP_FOLDER)
        except:
            try:
                ftp.mkd(FTP_FOLDER)
                ftp.cwd(FTP_FOLDER)
            except:
                pass  # Folder might already exist
        
        synced_files = []
        today = datetime.now().strftime("%Y-%m-%d")
        
        for user in users:
            user_id = user["id"]
            user_name = user.get("name", user.get("email", "unknown")).replace(" ", "_").replace("@", "_at_")
            
            # Get all workers for this user
            workers_cursor = db.workers.find({"owner_id": user_id}, {"_id": 0})
            workers = await workers_cursor.to_list(10000)
            
            if not workers:
                continue
            
            # Enrich workers with type names and project statuses
            for w in workers:
                # Get type name
                type_doc = await db.worker_types.find_one({"id": w.get("worker_type_id")}, {"_id": 0})
                w["worker_type_name"] = type_doc["name"] if type_doc else ""
                
                # Get project statuses
                pw_list = await db.project_workers.find({"worker_id": w["id"]}, {"_id": 0}).to_list(100)
                project_statuses = []
                for pw in pw_list:
                    project = await db.projects.find_one({"id": pw["project_id"]}, {"_id": 0})
                    status = await db.statuses.find_one({"id": pw.get("status_id")}, {"_id": 0})
                    if project:
                        project_statuses.append({
                            "project_name": project["name"],
                            "status_name": status["name"] if status else "Hozzárendelve"
                        })
                w["project_statuses"] = project_statuses
            
            # Generate Excel
            excel_data = generate_recruiter_excel(user_id, user_name, workers)
            
            # Upload to FTP
            filename = f"{user_name}_dolgozok_{today}.xlsx"
            ftp.storbinary(f"STOR {filename}", io.BytesIO(excel_data))
            synced_files.append(filename)
            logger.info(f"FTP: Uploaded {filename} with {len(workers)} workers")
        
        ftp.quit()
        
        return {
            "status": "success",
            "synced_files": synced_files,
            "date": today
        }
        
    except Exception as e:
        logger.error(f"FTP sync error: {str(e)}")
        return {"status": "error", "message": str(e)}

@api_router.post("/sync/ftp")
async def trigger_ftp_sync(user: dict = Depends(require_admin)):
    """Manually trigger FTP sync (admin only)"""
    result = await sync_to_ftp()
    return result

@api_router.get("/sync/status")
async def get_sync_status(user: dict = Depends(require_admin)):
    """Get FTP sync configuration status"""
    return {
        "ftp_configured": bool(FTP_HOST and FTP_USER and FTP_PASS),
        "ftp_host": FTP_HOST if FTP_HOST else "Not configured",
        "ftp_folder": FTP_FOLDER
    }

# ==================== HEALTH ====================

# ==================== NOTIFICATIONS ====================

@api_router.get("/notifications", response_model=List[NotificationResponse])
async def get_notifications(user: dict = Depends(get_current_user)):
    """Felhasználó értesítéseinek lekérése"""
    notifications = await db.notifications.find(
        {"user_id": user["id"]}, {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return [NotificationResponse(**n) for n in notifications]

@api_router.get("/notifications/unread-count")
async def get_unread_notification_count(user: dict = Depends(get_current_user)):
    """Olvasatlan értesítések száma"""
    count = await db.notifications.count_documents({
        "user_id": user["id"],
        "is_read": False
    })
    return {"count": count}

@api_router.put("/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str, user: dict = Depends(get_current_user)):
    """Értesítés olvasottként jelölése"""
    result = await db.notifications.update_one(
        {"id": notification_id, "user_id": user["id"]},
        {"$set": {"is_read": True}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Értesítés nem található")
    return {"message": "Megjelölve olvasottként"}

@api_router.put("/notifications/read-all")
async def mark_all_notifications_read(user: dict = Depends(get_current_user)):
    """Összes értesítés olvasottként jelölése"""
    await db.notifications.update_many(
        {"user_id": user["id"], "is_read": False},
        {"$set": {"is_read": True}}
    )
    return {"message": "Összes értesítés olvasottként jelölve"}

@api_router.delete("/notifications/{notification_id}")
async def delete_notification(notification_id: str, user: dict = Depends(get_current_user)):
    """Értesítés törlése"""
    result = await db.notifications.delete_one({
        "id": notification_id,
        "user_id": user["id"]
    })
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Értesítés nem található")
    return {"message": "Értesítés törölve"}

# ==================== CALENDAR ====================

@api_router.get("/calendar/trials")
async def get_calendar_trials(user: dict = Depends(get_current_user)):
    """Naptár események - próbák. Admin minden próbát lát, toborzó a sajátját."""
    trials = await db.trials.find({}, {"_id": 0}).to_list(1000)
    
    events = []
    for trial in trials:
        project = await db.projects.find_one({"id": trial["project_id"]}, {"_id": 0})
        if not project:
            continue
        
        # Check visibility for recruiters
        if user["role"] != "admin":
            # Toborzó csak azokat a próbákat látja, ahol van dolgozója
            trial_workers = await db.trial_workers.find({"trial_id": trial["id"]}, {"_id": 0}).to_list(100)
            worker_ids = [tw["worker_id"] for tw in trial_workers]
            
            # Check if any of these workers belong to the recruiter
            own_workers = await db.workers.count_documents({
                "id": {"$in": worker_ids},
                "owner_id": user["id"]
            })
            
            # Also check if recruiter is assigned to the project
            if own_workers == 0 and user["id"] not in project.get("recruiter_ids", []):
                continue
        
        # Get worker count for the trial
        trial_worker_count = await db.trial_workers.count_documents({"trial_id": trial["id"]})
        
        # Build event
        event_date = trial["date"]
        event_time = trial.get("time", "")
        
        # Create datetime string
        if event_time:
            start_datetime = f"{event_date}T{event_time}:00"
            # Assume 2 hour duration for trials
            end_hour = int(event_time.split(":")[0]) + 2
            end_time = f"{end_hour:02d}:{event_time.split(':')[1] if ':' in event_time else '00'}"
            end_datetime = f"{event_date}T{end_time}:00"
        else:
            start_datetime = f"{event_date}T09:00:00"
            end_datetime = f"{event_date}T11:00:00"
        
        events.append({
            "id": trial["id"],
            "title": f"{project['name']} - Próba",
            "start": start_datetime,
            "end": end_datetime,
            "project_id": project["id"],
            "project_name": project["name"],
            "project_client": project.get("client_name", ""),
            "location": project.get("location", ""),
            "training_location": project.get("training_location", ""),
            "notes": trial.get("notes", ""),
            "worker_count": trial_worker_count,
            "color": "#6366f1",  # Indigo
            "type": "trial"
        })
    
    return events

@api_router.get("/calendar/projects")
async def get_calendar_projects(user: dict = Depends(get_current_user)):
    """Projekt dátumok a naptárban"""
    if user["role"] == "admin":
        projects = await db.projects.find({}, {"_id": 0}).to_list(1000)
    else:
        # Toborzó csak hozzárendelt projekteket látja
        projects = await db.projects.find({
            "$or": [
                {"owner_id": user["id"]},
                {"recruiter_ids": user["id"]}
            ]
        }, {"_id": 0}).to_list(1000)
    
    events = []
    for p in projects:
        if not p.get("date"):
            continue
        
        worker_count = await db.project_workers.count_documents({"project_id": p["id"]})
        
        events.append({
            "id": p["id"],
            "title": p["name"],
            "start": f"{p['date']}T00:00:00",
            "end": f"{p['date']}T23:59:59",
            "allDay": True,
            "client_name": p.get("client_name", ""),
            "location": p.get("location", ""),
            "worker_count": worker_count,
            "is_closed": p.get("is_closed", False),
            "color": "#22c55e" if not p.get("is_closed") else "#94a3b8",  # Green or gray
            "type": "project"
        })
    
    return events

@api_router.get("/")
async def root():
    return {"message": "Dolgozó CRM API", "status": "running"}

@api_router.get("/health")
async def health():
    return {"status": "healthy"}

# ==================== APP CONFIG ====================

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
